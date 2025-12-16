from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404
from django.conf import settings
import os
from openpyxl import load_workbook
from .forms import BuyerForm, ClothesSearchForm, PurchaseForm
from .models import Buyer, Purchase, Assortment, AssortmentSize, ClothesType, Size, Seller
from django.db.models import Count, Sum, Q
from datetime import datetime, timedelta


def index(request):
    # Собираем статистику для главной страницы
    context = {
        'total_customers': Buyer.objects.count(),
        'total_purchases': Purchase.objects.count(),
        'total_revenue': Purchase.objects.aggregate(
            total=Sum('total_amount')
        )['total'] or 0,
        'total_items': Assortment.objects.count(),
    }
    return render(request, 'index.html', context)

def navigator(request):
    """Страница навигатора"""
    return render(request, 'navigator.html')


def buyers(request):
    """Список покупателей (простая версия)"""
    return HttpResponse("<h1>Список покупателей</h1><p>Здесь будет таблица покупателей.</p>")


def clothes_type(request):
    """Типы одежды"""
    return HttpResponse("<h1>Типы одежды</h1><p>Справочник: платья, брюки, футболки и т.д.</p>")


def assortment(request):
    """Ассортимент"""
    return HttpResponse("<h1>Ассортимент</h1><p>Список товаров с описанием и ценой.</p>")


def sizes(request):
    """Размеры"""
    return HttpResponse("<h1>Размеры</h1><p>Таблица размеров (S, M, L, XL, цифровые).</p>")


def sellers(request):
    """Продавцы"""
    return HttpResponse("<h1>Продавцы</h1><p>Информация о продавцах.</p>")


def purchases(request):
    """Покупки"""
    return HttpResponse("<h1>Покупки</h1><p>Журнал совершенных покупок.</p>")


def month_report(request, month):
    """Отчет по месяцам (re_path)"""
    months = {
        '01': 'Январь', '02': 'Февраль', '03': 'Март',
        '04': 'Апрель', '05': 'Май', '06': 'Июнь',
        '07': 'Июль', '08': 'Август', '09': 'Сентябрь',
        '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'
    }
    month_name = months.get(month, 'неизвестный месяц')
    return HttpResponse(f"<h1>Отчет за {month_name}</h1><p>Здесь будут данные о покупках.</p>")


# ============================================================================
# ПАРАМЕТРЫ В URL (Задача 1.1 продолжение)
# ============================================================================

def buyer_detail(request, buyer_id):
    """Детальная информация о покупателе (с параметром)"""
    context = {
        'buyer_id': buyer_id,
        'message': f'Вы просматриваете профиль покупателя с ID={buyer_id}.'
    }
    return render(request, 'buyer_detail.html', context)


def clothes_list(request, category='all'):
    """Список одежды по категориям (параметр по умолчанию)"""
    clothes_data = {
        'all': ['Платья', 'Брюки', 'Футболки', 'Куртки', 'Юбки', 'Джинсы'],
        'outerwear': ['Куртки', 'Пальто', 'Плащи', 'Ветровки'],
        'underwear': ['Футболки', 'Майки', 'Боди', 'Топики'],
        'dress': ['Вечерние платья', 'Коктейльные платья', 'Повседневные платья'],
        'pants': ['Джинсы', 'Брюки', 'Шорты', 'Леггинсы']
    }
    items = clothes_data.get(category, ['Категория не найдена'])

    context = {
        'category': category,
        'items': items
    }
    return render(request, 'clothes_list.html', context)


def purchase_year(request, year):
    """Покупки за год (re_path с регулярным выражением)"""
    return HttpResponse(f'<h1>Покупки за {year} год</h1><p>Здесь будет статистика за выбранный год.</p>')


def search_assortment(request):
    """Поиск ассортимента с параметрами строки запроса"""
    name = request.GET.get('name', '')
    min_price = request.GET.get('min_price', '0')
    max_price = request.GET.get('max_price', '100000')

    context = {
        'search_name': name,
        'min_price': min_price,
        'max_price': max_price,
        'message': f'Поиск: "{name}", цена от {min_price} до {max_price} руб.'
    }
    return render(request, 'search_result.html', context)


def old_purchases_page(request):
    """Страница с переадресацией"""
    return redirect('purchases')


# ============================================================================
# ДАННЫЕ ИЗ ФАЙЛОВ (Задача 1.2)
# ============================================================================

def show_tablica_data(request):
    """Отображение данных из текстовых и Excel файлов"""
    txt_file_path = os.path.join(settings.BASE_DIR, 'firstapp_var_11', 'tablica', 'buyers.txt')
    excel_file_path = os.path.join(settings.BASE_DIR, 'firstapp_var_11', 'tablica', 'clothes.xlsx')

    # Чтение из текстового файла
    buyers_list = []
    if os.path.exists(txt_file_path):
        with open(txt_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    parts = line.strip().split(';')
                    buyers_list.append({
                        'name': parts[0],
                        'email': parts[1] if len(parts) > 1 else '',
                        'city': parts[2] if len(parts) > 2 else ''
                    })

    # Чтение из Excel файла
    clothes_table = []
    headers = []
    if os.path.exists(excel_file_path):
        try:
            wb = load_workbook(excel_file_path)
            ws = wb.active

            # Читаем заголовки
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Читаем данные
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value
                    clothes_table.append(row_dict)
        except Exception as e:
            clothes_table = [{'Ошибка': f'Не удалось прочитать Excel файл: {str(e)}'}]

    # Примеры данных для передачи в шаблон
    simple_data = {
        'string': 'Пример текстовой строки',
        'number': 42,
        'float_number': 3.14159,
        'boolean': True,
        'date': '2024-01-15',
        'null_value': None
    }

    complex_data = {
        'list': ['Платье вечернее', 'Джинсы скинни', 'Футболка поло', 'Куртка кожаная'],
        'tuple': ('XS', 'S', 'M', 'L', 'XL'),
        'dict': {'цена': 2500, 'количество': 42, 'скидка': 15},
        'list_of_dicts': [
            {'id': 1, 'название': 'Платье', 'цена': 1500},
            {'id': 2, 'название': 'Брюки', 'цена': 2000},
            {'id': 3, 'название': 'Рубашка', 'цена': 1200}
        ]
    }

    context = {
        'buyers': buyers_list,
        'clothes': clothes_table,
        'simple': simple_data,
        'complex': complex_data,
        'headers': headers if headers else ['Нет данных'],
        'has_excel': os.path.exists(excel_file_path),
        'has_txt': os.path.exists(txt_file_path)
    }

    return render(request, 'tablica_data.html', context)


def assortment_gallery(request):
    """Галерея товаров с изображениями"""
    gallery_items = [
        {'name': 'Платья', 'image': 'dress.jpg', 'count': 42, 'description': 'Вечерние и повседневные платья'},
        {'name': 'Брюки', 'image': 'pants.jpg', 'count': 35, 'description': 'Джинсы, классические брюки'},
        {'name': 'Рубашки', 'image': 'shirt.jpg', 'count': 28, 'description': 'Мужские и женские рубашки'},
        {'name': 'Куртки', 'image': 'jacket.jpg', 'count': 19, 'description': 'Кожаные и джинсовые куртки'},
        {'name': 'Юбки', 'image': 'skirt.jpg', 'count': 24, 'description': 'Различные фасоны юбок'},
        {'name': 'Футболки', 'image': 'tshirt.jpg', 'count': 57, 'description': 'Базовые и принтованные'},
    ]

    return render(request, 'assortment_gallery.html', {'gallery': gallery_items})


# ============================================================================
# ФОРМЫ (Задача 1.3)
# ============================================================================

def buyer_form_view(request):
    """Форма для добавления покупателя (POST)"""
    if request.method == 'POST':
        form = BuyerForm(request.POST)
        if form.is_valid():
            buyer_data = form.cleaned_data
            # Сохранение в файл (для демонстрации)
            txt_file_path = os.path.join(settings.BASE_DIR, 'firstapp_var_11', 'tablica', 'buyers_new.txt')
            with open(txt_file_path, 'a', encoding='utf-8') as f:
                f.write(f"{buyer_data['name']};{buyer_data['email']};{buyer_data['city']}\n")

            return render(request, 'form_success.html', {'buyer_data': buyer_data})
    else:
        form = BuyerForm()

    return render(request, 'buyer_form.html', {'form': form})


def clothes_search_view(request):
    """Форма поиска одежды (GET с выбором)"""
    if request.method == 'GET':
        form = ClothesSearchForm(request.GET)
        if form.is_valid():
            search_data = form.cleaned_data
            return render(request, 'search_results.html', {'form': form, 'search_data': search_data})
    else:
        form = ClothesSearchForm()

    return render(request, 'clothes_search.html', {'form': form})


def purchase_form_view(request):
    """Форма оформления покупки"""
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            return render(request, 'purchase_success.html', {'purchase_data': data})
    else:
        form = PurchaseForm()

    return render(request, 'purchase_form.html', {'form': form})


# ============================================================================
# ДОПОЛНИТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================================

def buyer_purchases(request, buyer_id):
    """Демонстрация связи 1:M"""
    try:
        buyer = Buyer.objects.get(id=buyer_id)
        purchases = Purchase.objects.filter(buyer=buyer)
        context = {
            'buyer': buyer,
            'purchases': purchases,
            'count': purchases.count()
        }
        return render(request, 'buyer_purchases.html', context)
    except Buyer.DoesNotExist:
        raise Http404("Покупатель не найден")


def clothes_sizes(request, clothes_id):
    """Демонстрация связи M:M"""
    try:
        clothes = Assortment.objects.get(id=clothes_id)
        sizes = clothes.available_sizes.all()
        context = {
            'clothes': clothes,
            'sizes': sizes,
            'count': sizes.count()
        }
        return render(request, 'clothes_sizes.html', context)
    except Assortment.DoesNotExist:
        raise Http404("Товар не найден")


def handler404(request, exception):
    """Обработчик 404 ошибки"""
    return render(request, '404.html', status=404)


def handler500(request):
    """Обработчик 500 ошибки"""
    return render(request, '500.html', status=500)


# ============================================================================
# АНАЛИТИЧЕСКИЕ ФУНКЦИИ
# ============================================================================

def sales_stats(request):
    """Общая статистика продаж"""
    from django.db.models import Sum, Avg, Count
    from datetime import datetime, timedelta
    import json
    
    # #region agent log
    log_path = r'c:\Users\kiril\PycharmProjects\PythonProject_Kursovaya\.cursor\debug.log'
    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(json.dumps({'location': 'views.py:317', 'message': 'sales_stats entry', 'data': {'hypothesisId': 'C'}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
    except: pass
    # #endregion
    
    try:
        total_revenue = Purchase.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        avg_purchase = Purchase.objects.aggregate(Avg('total_amount'))['total_amount__avg'] or 0
        total_purchases = Purchase.objects.count()
        
        # Статистика за последние 30 дней
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_purchases = Purchase.objects.filter(purchase_date__gte=thirty_days_ago)
        recent_revenue = recent_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:330', 'message': 'sales_stats calculations successful', 'data': {'hypothesisId': 'C', 'total_revenue': float(total_revenue), 'total_purchases': total_purchases}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
    except Exception as e:
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:332', 'message': 'sales_stats failed', 'data': {'hypothesisId': 'C', 'error': str(e), 'error_type': type(e).__name__}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
        raise
    
    context = {
        'total_revenue': total_revenue,
        'avg_purchase': avg_purchase,
        'total_purchases': total_purchases,
        'recent_revenue': recent_revenue,
        'recent_purchases_count': recent_purchases.count(),
    }
    return render(request, 'analytics/sales_stats.html', context)


def sales_by_customer(request):
    """Продажи по покупателям"""
    from django.db.models import Sum, Count
    
    customers_stats = Buyer.objects.annotate(
        total_purchases=Count('purchases'),
        total_spent=Sum('purchases__total_amount')
    ).order_by('-total_spent')
    
    context = {
        'customers_stats': customers_stats,
    }
    return render(request, 'analytics/sales_by_customer.html', context)


def sales_by_type(request):
    """Продажи по типам одежды"""
    from django.db.models import Sum, Count
    
    # Здесь нужно будет добавить связь Purchase с Assortment через промежуточную модель
    # Пока используем данные из Assortment
    type_stats = ClothesType.objects.annotate(
        items_count=Count('assortments'),
        total_value=Sum('assortments__price')
    ).order_by('-items_count')
    
    context = {
        'type_stats': type_stats,
    }
    return render(request, 'analytics/sales_by_type.html', context)


def assortment_by_size(request):
    """Ассортимент по размерам"""
    from django.db.models import Count, Sum
    import json
    import os
    
    # #region agent log
    log_path = r'c:\Users\kiril\PycharmProjects\PythonProject_Kursovaya\.cursor\debug.log'
    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(json.dumps({'location': 'views.py:373', 'message': 'assortment_by_size entry', 'data': {'hypothesisId': 'A'}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
    except: pass
    # #endregion
    
    try:
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:378', 'message': 'Before query execution', 'data': {'hypothesisId': 'A', 'query_field': 'assortmentsize__assortment'}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
        
        size_stats = Size.objects.annotate(
            items_count=Count('assortmentsize__assortment'),
            total_quantity=Sum('assortmentsize__quantity')
        ).order_by('-items_count')
        
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:385', 'message': 'Query executed successfully', 'data': {'hypothesisId': 'A', 'result_count': len(list(size_stats))}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
    except Exception as e:
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:387', 'message': 'Query failed', 'data': {'hypothesisId': 'A', 'error': str(e), 'error_type': type(e).__name__}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
        raise
    
    context = {
        'size_stats': size_stats,
    }
    return render(request, 'analytics/assortment_by_size.html', context)


def seller_performance(request):
    """Эффективность продавцов"""
    import json
    import os
    
    # #region agent log
    log_path = r'c:\Users\kiril\PycharmProjects\PythonProject_Kursovaya\.cursor\debug.log'
    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(json.dumps({'location': 'views.py:388', 'message': 'seller_performance entry', 'data': {'hypothesisId': 'B'}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
    except: pass
    # #endregion
    
    # Пока продавцы не связаны с покупками, возвращаем базовую информацию
    try:
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:391', 'message': 'Before Seller.objects.all()', 'data': {'hypothesisId': 'B'}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
        
        sellers = Seller.objects.all()
        
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:393', 'message': 'Seller query successful', 'data': {'hypothesisId': 'B', 'sellers_count': sellers.count()}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
    except Exception as e:
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({'location': 'views.py:395', 'message': 'Seller query failed', 'data': {'hypothesisId': 'B', 'error': str(e), 'error_type': type(e).__name__}, 'timestamp': __import__('time').time() * 1000, 'sessionId': 'debug-session', 'runId': 'run1'}) + '\n')
        except: pass
        # #endregion
        raise
    
    context = {
        'sellers': sellers,
    }
    return render(request, 'analytics/seller_performance.html', context)


def daily_sales_report(request):
    """Ежедневный отчет по продажам"""
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    
    daily_stats = []
    for i in range(7):
        date = week_ago + timedelta(days=i)
        purchases = Purchase.objects.filter(purchase_date__date=date)
        daily_stats.append({
            'date': date,
            'count': purchases.count(),
            'revenue': purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        })
    
    context = {
        'daily_stats': daily_stats,
    }
    return render(request, 'analytics/daily_sales.html', context)


def customer_segments(request):
    """Сегментация покупателей"""
    from django.db.models import Sum, Count, Avg
    
    vip_customers = Buyer.objects.filter(is_vip=True).count()
    regular_customers = Buyer.objects.filter(is_vip=False).count()
    
    customers_with_purchases = Buyer.objects.annotate(
        purchase_count=Count('purchases'),
        total_spent=Sum('purchases__total_amount')
    ).filter(purchase_count__gt=0)
    
    high_value = customers_with_purchases.filter(total_spent__gte=10000).count()
    medium_value = customers_with_purchases.filter(
        total_spent__gte=5000, total_spent__lt=10000
    ).count()
    low_value = customers_with_purchases.filter(total_spent__lt=5000).count()
    
    context = {
        'vip_customers': vip_customers,
        'regular_customers': regular_customers,
        'high_value': high_value,
        'medium_value': medium_value,
        'low_value': low_value,
    }
    return render(request, 'analytics/customer_segments.html', context)


def inventory_status(request):
    """Статус склада"""
    from django.db.models import Sum, Count
    
    total_items = Assortment.objects.count()
    in_stock = Assortment.objects.filter(stock_quantity__gt=0).count()
    out_of_stock = Assortment.objects.filter(stock_quantity=0).count()
    low_stock = Assortment.objects.filter(stock_quantity__lte=5, stock_quantity__gt=0).count()
    
    total_value = Assortment.objects.aggregate(
        total=Sum('price')
    )['total'] or 0
    
    context = {
        'total_items': total_items,
        'in_stock': in_stock,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'total_value': total_value,
    }
    return render(request, 'analytics/inventory_status.html', context)


def top_products(request):
    """Топ товаров"""
    from django.db.models import Count
    
    # Пока просто сортируем по количеству на складе
    top_products = Assortment.objects.order_by('-stock_quantity')[:10]
    
    context = {
        'top_products': top_products,
    }
    return render(request, 'analytics/top_products.html', context)


def purchase_trends(request):
    """Тренды покупок"""
    from django.db.models import Count
    from django.utils import timezone
    from datetime import timedelta
    from collections import defaultdict
    
    # Статистика по способам оплаты
    payment_stats = Purchase.objects.values('payment_method').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Статистика по месяцам
    monthly_stats = defaultdict(int)
    purchases = Purchase.objects.all()
    for purchase in purchases:
        month_key = purchase.purchase_date.strftime('%Y-%m')
        monthly_stats[month_key] += 1
    
    context = {
        'payment_stats': payment_stats,
        'monthly_stats': dict(sorted(monthly_stats.items())),
    }
    return render(request, 'analytics/purchase_trends.html', context)